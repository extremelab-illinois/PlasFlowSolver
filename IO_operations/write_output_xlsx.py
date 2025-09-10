# .................................................
#   WRITE_OUTPUT_XLSX.PY, v2.0.0, December 2024, Domenico Lanza.
# .................................................
#   This module is needed to write the output file
#   in xlsx format.
# .................................................
import pandas as pd  # Library to manage dataframes
# Library to manage excel files
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Alignment

def write_output_xlsx(output_filename, out_obj):
    """This function writes the output file in xlsx format.

    Args:
        output_filename (str): the name of the output file
        out_obj (out_properties_class): the object containing all the output properties
    """
    # Extracting the output properties:
    has_converged_out = out_obj.has_converged_out  # Has converged flag
    rho_out = out_obj.rho_out  # Density
    T_out = out_obj.T_out  # Static temperature
    h_out = out_obj.h_out  # Static enthalpy
    u_out = out_obj.u_out  # Flow velocity
    a_out = out_obj.a_out  # Speed of sound
    M_out = out_obj.M_out  # Mach number
    T_t_out = out_obj.T_t_out  # Total temperature
    h_t_out = out_obj.h_t_out  # Total enthalpy
    P_t_out = out_obj.P_t_out  # Total pressure
    Re_out = out_obj.Re_out  # Reynolds number
    Kn_out = out_obj.Kn_out  # Knudsen number
    warnings_out = out_obj.warnings_out  # Warnings
    res_out = out_obj.res_out  # Final convergence criteria
    species_names_out = out_obj.species_names_out  # Names of the species (dictionary)
    species_Y_out = out_obj.species_Y_out  # Mass fractions of the species (dictionary)

    # Rebuild the input file name:
    input_filename = output_filename[:-9] + ".xlsx"

    # Read the dataframe from the input file:
    df = pd.read_excel(input_filename, header=[0, 1])

    # Scale the output data:
    for i in range(0, len(rho_out)):
        if (rho_out[i] != -1):
            rho_out[i] = rho_out[i] * 1000  # From kg/m^3 to g/m^3
            h_out[i] = h_out[i] / 1000  # From J/kg to kJ/kg
            h_t_out[i] = h_t_out[i] / 1000  # From J/kg to kJ/kg
            P_t_out[i] = P_t_out[i] / 1000  # From Pa to kPa

    # Add the new columns to the dataframe:
    # Has converged
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "has converged"), has_converged_out, False)
    
    # Residuals
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "residual"), res_out, False)

    # Density
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "density [g/m^3]"), rho_out, False)

    # Temperature
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "temperature [K]"), T_out, False)

    # Enthalpy
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "enthalpy [kJ/kg]"), h_out, False)

    # Velocity
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "velocity [m/s]"), u_out, False)

    # Speed of sound
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "speed of sound [m/s]"), a_out, False)

    # Mach number
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "mach number"), M_out, False)

    # Total temperature
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "total temperature [K]"), T_t_out, False)

    # Total enthalpy
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "total enthalpy [kJ/kg]"), h_t_out, False)

    # Total pressure
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "total pressure [kPa]"), P_t_out, False)

    # Reynolds number
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "reynolds number"), Re_out, False)

    # Knudsen number
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "knudsen number"), Kn_out, False)

    # Species names and mass fractions
    n_cases = len(species_names_out)
    if len(species_Y_out) != n_cases:
        raise ValueError("The number of cases in the species names and mass fractions dictionaries is different. This should not be possible.")
    for i in range(1, n_cases + 1):  # From 1 to n_cases included
        c_species_names = species_names_out[i]  # I extract the current species names
        c_species_Y = species_Y_out[i]  # I extract the current species mass fractions
        if (c_species_names is None) or (c_species_Y is None):
            continue
        for j in range(len(c_species_names)):
            c_species_name = c_species_names[j]
            if ("Output", c_species_name) not in df.columns:
                n_col = len(df.columns)
                df.insert(n_col, ("Output", c_species_name), pd.NA, False)
        for j in range(len(c_species_names)):
            c_species_name = c_species_names[j]
            c_species_Y_value = c_species_Y[j]
            df.loc[i - 1, ("Output", c_species_name)] = c_species_Y_value

    # Warnings
    n_col = len(df.columns)
    df.insert(n_col, ("Output", "warnings"), warnings_out, False)

    # Replace <NA> values with empty strings
    with pd.option_context("future.no_silent_downcasting", True):
        df = df.fillna('').infer_objects(copy=False)

    # Opening an excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    
    # Writing dataframe to excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Merging header cells
    # Inputs: merge the first 6 cells
    ws.merge_cells(start_row=1, end_row=1, start_column=1, end_column=6)
    # Initial conditions: merge the next 5 cells
    ws.merge_cells(start_row=1, end_row=1, start_column=7, end_column=11)
    # Probe settings: merge the next 7 cells
    ws.merge_cells(start_row=1, end_row=1, start_column=12, end_column=18)
    # Program settings: merge the next 11 cells
    ws.merge_cells(start_row=1, end_row=1, start_column=19, end_column=29)
    # Outputs: merge the rest of the cells to the end
    ws.merge_cells(start_row=1, end_row=1, start_column=30, end_column=df.shape[1])
    # Set the first 2 rows in bold
    for cell in ws["1:2"]:
        for c in cell:
            c.font = c.font.copy(bold=True)
    # Align the text to the center for all cells
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
    # Adjust the column width to fit the text in each cell
    for col in ws.iter_cols(min_row=2, max_row=df.shape[0]+2, min_col=1, max_col=df.shape[1]):
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
    # Color the first 2 rows and 6 columns in #FFFF00
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=1, max_col=6):
        for cell in row:
            cell.fill = yellow_fill
    # Color the first 2 rows and 5 columns in #F6C6AD
    pink_fill = PatternFill(start_color="F6C6AD", end_color="F6C6AD", fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=7, max_col=11):
        for cell in row:
            cell.fill = pink_fill
    # Color the first 2 rows and 7 columns in #96DCF8
    blue_fill = PatternFill(start_color="96DCF8", end_color="96DCF8", fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=12, max_col=18):
        for cell in row:
            cell.fill = blue_fill
    # Color the first 2 rows and 11 columns in #E59EDD
    purple_fill = PatternFill(start_color="E59EDD", end_color="E59EDD", fill_type="solid")
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=19, max_col=29):
        for cell in row:
            cell.fill = purple_fill
    # Color the first 2 rows and the rest of the columns in #C4D79B
    for row in ws.iter_rows(min_row=1, max_row=2, min_col=30, max_col=df.shape[1]):
        for cell in row:
            cell.fill = PatternFill(start_color="C4D79B", end_color="C4D79B", fill_type="solid")
    # Add all borders
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.border = border
    # Save to excel
    wb.save(output_filename)
#.................................................
#   Possible improvements:
#   None.
#.................................................
#   Known problems:
#   None.
#.................................................